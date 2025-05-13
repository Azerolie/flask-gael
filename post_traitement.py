import openpyxl
from openpyxl.styles import Font
import os

# Fonction pour nettoyer les feuilles d'importation (avec "imp_" dans le nom de la feuille)
def clean_import_sheets(input_file, output_file):
    if not os.path.exists(input_file):
        print(f"❌ Le fichier '{input_file}' n'existe pas.")
        return

    try:
        # Charger le fichier modifié par l'utilisateur
        wb = openpyxl.load_workbook(input_file, data_only=True)  # data_only=True pour avoir les valeurs des formules
        new_wb = openpyxl.Workbook()
        new_wb.remove(new_wb.active)  # Supprimer la feuille par défaut

        for sheet_name in wb.sheetnames:
            if sheet_name.startswith("imp_"):
                print(f"🔍 Traitement de la feuille : {sheet_name}")
                ws = wb[sheet_name]
                new_ws = new_wb.create_sheet(title=sheet_name)

                # Ajouter les en-têtes
                headers = ["Entité", "Champs", "Valeur externe", "Valeur"]
                new_ws.append(headers)
                for cell in new_ws[1]:
                    cell.font = Font(bold=True)

                # Traiter chaque ligne à partir de la 2e (en-têtes ignorés)
                for row in ws.iter_rows(min_row=2):
                    code_client = row[2].value  # Colonne C (index 2)
                    id_ximi = row[5].value      # Colonne F (index 5)

                    if code_client and id_ximi:
                        print(f"✅ Ajout : {code_client} → {id_ximi}")
                        new_ws.append(["Amendment", "PayJobMark", code_client, id_ximi])

        # Sauvegarder le fichier nettoyé
        new_wb.save(output_file)
        print(f"✅ Fichier '{output_file}' généré avec succès.")

    except Exception as e:
        print(f"❌ Erreur pendant le traitement : {e}")