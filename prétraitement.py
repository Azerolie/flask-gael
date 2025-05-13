import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
import re

def protect_validation_sheet(ws):
    for row in ws.iter_rows():
        for cell in row:
            cell.protection = openpyxl.styles.Protection(locked=True)
    
    for validation in ws.data_validations:
        if hasattr(validation, 'ranges'):
            for cell_range in validation.ranges:
                for row in range(cell_range.min_row, cell_range.max_row + 1):
                    for col in range(cell_range.min_col, cell_range.max_col + 1):
                        ws.cell(row=row, column=col).protection = openpyxl.styles.Protection(locked=False)
    
    ws.protection.sheet = True

def parse_mapping_file(input_file):
    mappings = []
    with open(input_file, "r", encoding="utf-8") as file:
        for line in file:
            line = line.strip()
            if not line:
                continue
            if "," in line and "->" not in line and "{" not in line:
                parts = line.split(",")
                if len(parts) == 2:
                    mappings.append((parts[1].strip(), parts[0].strip()))
            elif "->" in line:
                parts = line.split("->")
                if len(parts) == 2:
                    mappings.append((parts[1].strip(), parts[0].strip()))
            elif "{" in line and "}" in line:
                matches = re.findall(r'\{?\s*"([^"]*)"\s*,\s*([^\}]*)\s*\}', line)
                if matches:
                    code, libelle = matches[0]
                    mappings.append((libelle.strip(), code.strip()))
    return mappings

def load_excel_data(file_path):
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        return [(row[0], row[1]) for row in ws.iter_rows(min_row=2, max_col=2, values_only=True) if row[0]]
    except Exception as e:
        print(f"Erreur lors du chargement de {file_path} : {e}")
        return []

def create_validations_sheet(wb, libelle_ximi_data, code_ximi_list, objet_ximi_list, tab_type):
    ws = wb.create_sheet(title=f"Validations_{tab_type}")
    headers = {
        "import": ["Libellé Ximi", "ID Ximi"],
        "evp": ["Code Ximi"],
        "export": ["Objet Ximi"]
    }
    ws.append(headers[tab_type])
    data_map = {
        "import": libelle_ximi_data,
        "evp": code_ximi_list,
        "export": [(o,) for o in objet_ximi_list]
    }
    for row in data_map[tab_type]:
        ws.append(row)
    protect_validation_sheet(ws)

def create_import_sheet(wb, tab_name):
    ws = wb.create_sheet(title=f"imp_{tab_name}")
    ws.append(["Convention collective", "Type contrat", "Code client", "Libellé client", "Libellé Ximi", "ID Ximi", "Commentaire"])
    validation_range = "Validations_import!$A$2:$A$300"
    dv = DataValidation(type="list", formula1=validation_range, allow_blank=False)
    ws.add_data_validation(dv)
    for row in ws.iter_rows(min_row=2, max_row=101, min_col=5, max_col=5):
        for cell in row:
            dv.add(cell)
    for row in range(2, ws.max_row + 1):  
        ws[f"F{row}"].value = f'=IF(E{row}="","",VLOOKUP(E{row}, Validations_import!$A$2:$B$300, 2, FALSE))'

def create_evp_sheet(wb, tab_name):
    ws = wb.create_sheet(title=f"evp_{tab_name}")
    ws.append(["Evenement client", "Code client", "Libellé Ximi", "Code Ximi", "Commentaire"])
    validation_range = "Validations_evp!A$2:$A$300"
    dv = DataValidation(type="list", formula1=validation_range, allow_blank=False)
    ws.add_data_validation(dv)
    for row in ws.iter_rows(min_row=2, max_row=101, min_col=3, max_col=3):
        for cell in row:
            dv.add(cell)
    for row in range(2, ws.max_row + 1):
        ws[f"D{row}"].value = f'=IF(C{row}="","",VLOOKUP(C{row}, Validations_evp!$B$2:$A$300, 2, FALSE))'

def create_export_sheet(wb, tab_name):
    ws = wb.create_sheet(title=f"export_{tab_name}")
    ws.append(["Objet Ximi", "Code Ximi", "Commentaire"])
    validation_range = "Validations_export!$A$2:$A$300"
    dv = DataValidation(type="list", formula1=validation_range, allow_blank=False)
    ws.add_data_validation(dv)
    for row in ws.iter_rows(min_row=2, max_row=101, min_col=1, max_col=1):
        for cell in row:
            dv.add(cell)

def prétraitement(tabs):
    wb = Workbook()
    wb.remove(wb.active)

    libelle_ximi_data = load_excel_data("Export_Emplois_Repères_(Générale).xlsx")
    code_ximi_list = parse_mapping_file("mapping.txt")
    objet_ximi_list = [row[0] for row in load_excel_data("objet_ximi.xlsx")]

    for tab in tabs:
        tab_name = tab["nom"]
        tab_type = tab["type"].lower()

        create_validations_sheet(wb, libelle_ximi_data, code_ximi_list, objet_ximi_list, tab_type)

        if tab_type == "import":
            create_import_sheet(wb, tab_name)
        elif tab_type == "evp":
            create_evp_sheet(wb, tab_name)
        elif tab_type == "export":
            create_export_sheet(wb, tab_name)
        else:
            print(f"❌ Type d'onglet inconnu : {tab_type}")

    wb.security.workbookPassword = "securepassword"
    wb.security.lockStructure = True
    wb.save("mon_fichier.xlsx")